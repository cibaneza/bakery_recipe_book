from django.shortcuts import render
from django.shortcuts import redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from .models import Product, Ingredient, ProductIngredient
from .forms import ProductForm, IngredientForm
from django.views.decorators.csrf import csrf_exempt
import json
from openpyxl import Workbook
#from openpyxl.writer.excel import save_virtual_workbook
from tempfile import NamedTemporaryFile
from io import BytesIO
from django.core import serializers

def home(request):
    context = { 
               "current_tab": "home", 
               }
    return render(request, 'home.html', context=context)

def ingredients(request):
    
    ingredient = Ingredient.objects.all()
    form = IngredientForm()
    
    if request.method == "POST":
        form = IngredientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('ingredients')
    
    context = { 
               "current_tab": "ingredients",
               "ingredients": ingredient,
               "form": form,
               }
    
    return render(request, 'ingredients.html', context=context)

def receiptcosts(request):
    product = Product.objects.all()
    form = ProductForm()
    
    if request.method == "POST":
        form = ProductForm(request.POST)
        if form.is_valid(): 
            form.save()
            return redirect('receipt-costs')
     
    context = { 
               "current_tab": "receipt-costs",
               "products": product,
               "form": form,
               }
    return render(request, 'receipt_costs.html', context=context)


def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    product_ingredients = product.productingredient_set.all()

    context = {
        "product": product,
        "product_ingredients": product_ingredients,
    }
    return render(request, 'product_detail.html', context)

def delete_product_ingredient(request):
    product_ingredient_id = request.POST.get('productIngredientId')
    try:
        product_ingredient = ProductIngredient.objects.get(id=product_ingredient_id)
        product_ingredient.delete()
        return JsonResponse({'success': True})
    except ProductIngredient.DoesNotExist:
        return JsonResponse({'success': False}, status=404)

def get_product_ingredient_data(request, ingredient_id):
    try:
        ingredient = ProductIngredient.objects.get(id=ingredient_id)
        data = {
            'id': ingredient.id,
            'quantity': ingredient.quantity,
            'cost': ingredient.cost,
            'unit': ingredient.unit,
        }
        return JsonResponse(data)
    except ProductIngredient.DoesNotExist:
        return JsonResponse({'error': 'Ingrediente no encontrado'}, status=404)

@require_POST
def update_product_ingredient(request):
    data = json.loads(request.body)
    product_ingredient_id = data.get('productIngredientId')
    quantity = data.get('quantity')
    cost = data.get('cost')
    unit = data.get('unit')
    
    if not unit:
        return JsonResponse({'error': 'El campo unit es obligatorio'}, status=400)
    
    try:
        product_ingredient = ProductIngredient.objects.get(id=product_ingredient_id)
        product_ingredient.quantity = quantity
        product_ingredient.cost = cost
        product_ingredient.unit = unit
        product_ingredient.save()
        return JsonResponse({'success': True})
    except ProductIngredient.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Ingrediente no encontrado'}, status=404)


@require_POST
@csrf_exempt  
def add_product_ingredient(request):
    
    if request.headers.get('Content-Type') == 'application/json':
        data = json.loads(request.body)
        product_id = data.get('productId')
        ingredient_id = data.get('ingredientId')

        try:
            product = Product.objects.get(id=product_id)
            ingredient = Ingredient.objects.get(id=ingredient_id)
            product_ingredient, created = ProductIngredient.objects.get_or_create(
                product=product, 
                ingredient=ingredient,
                defaults={'quantity': data.get('quantity', 0), 'cost': data.get('cost', 0)}  
            )

            if not created:
                if 'quantity' in data:
                    product_ingredient.quantity = data['quantity']
                if 'cost' in data:
                    product_ingredient.cost = data['cost']
                product_ingredient.save()

            return JsonResponse({'status': 'success', 'message': 'Ingrediente añadido o actualizado correctamente.'})
        except Product.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'El producto no existe.'}, status=400)
        except Ingredient.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'El ingrediente no existe.'}, status=400)
        except Exception as e:
            
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    else:
        return JsonResponse({'status': 'error', 'message': 'Tipo de contenido incorrecto.'}, status=400)

def get_ingredients(request):
    ingredients = Ingredient.objects.all().values('id', 'name')
    return JsonResponse(list(ingredients), safe=False)

def get_ingredient_units(request):
    # units = ProductIngredient.objects.all().values('id', 'unit')
    # units = ProductIngredient.objects.values_list('unit', flat=True).distinct()
    # units = list(filter(None, units))  # Elimina valores None o vacíos.
    units = [unit[1] for unit in ProductIngredient.unit.field.choices]
    return JsonResponse({'units': units})
    #return JsonResponse({'units': list(units)})

def download_product_data_excel(request):
    # Crea un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Añade los encabezados de columna en la primera fila
    columns = ['Id', 'Producto', 'Descripción', 'Rendimiento', 'Costo Total', 'Costo/Unidad']
    ws.append(columns)

    # Recupera los productos y sus datos para añadirlos al Excel
    products = Product.objects.all()
    for product in products:
        ws.append([
            product.id, 
            product.name, 
            product.description, 
            product.yield_quantity, 
            product.get_total_cost(), 
            product.get_cost_per_unit()
        ])

    # Guarda el libro de trabajo en un objeto BytesIO
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="productos-Bakery-Merry.xlsx"'
    with BytesIO() as b:
        wb.save(b)  # Guarda el libro de trabajo en el objeto BytesIO
        response.write(b.getvalue())  # Escribe el contenido del BytesIO en la respuesta HTTP

    return response

def download_ingredients_data_excel(request):
    # Crea un libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Añade los encabezados de columna en la primera fila
    columns = ['Id', 'Ingrediente', 'Descripción', 'Fecha de Creación']
    ws.append(columns)

    # Recupera los productos y sus datos para añadirlos al Excel
    # products = Product.objects.all()
    products = Ingredient.objects.all()
    for product in products:
        ws.append([
            product.id, 
            product.name, 
            product.description, 
            product.date
        ])

    # Guarda el libro de trabajo en un objeto BytesIO
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ingredientes-Bakery-Merry.xlsx"'
    with BytesIO() as b:
        wb.save(b)  # Guarda el libro de trabajo en el objeto BytesIO
        response.write(b.getvalue())  # Escribe el contenido del BytesIO en la respuesta HTTP

    return response

def obtener_datos_recetario(request):
    # Lista para almacenar todos los productos y sus ingredientes
    products_list = []

    # Obtener todos los productos
    products = Product.objects.all()

    for product in products:
        # Diccionario para almacenar información del producto
        product_dict = {
            "name": product.name,
            "description": product.description,
            "date": product.date,
            "yield_quantity": product.yield_quantity,
            "total_cost": product.get_total_cost(),
            "cost_per_unit": product.get_cost_per_unit(),
            "ingredients": []
        }

        # Obtener todos los ingredientes para este producto
        product_ingredients = product.productingredient_set.all()

        for pi in product_ingredients:
            product_dict["ingredients"].append({
                "name": pi.ingredient.name,
                "description": pi.ingredient.description,
                "cost": pi.cost,
                "quantity": pi.quantity,
                "unit": pi.unit
            })

        products_list.append(product_dict)

    # Devolver los datos serializados como JSON
    # return JsonResponse(products_list, safe=False)
    return products_list 

def download_products_with_ingredients(request):
    format = request.GET.get('format', 'json')
    datos = obtener_datos_recetario(request)

    if format == 'excel' or format == 'xlsx' or format == 'xls':
        wb = Workbook()
        ws = wb.active
        ws.append(['Nombre', 'Descripción', 'Fecha', 'Rendimiento', 'Costo Total', 'Costo por Unidad', 'Ingredientes'])
        
        for producto in datos:
            # Añade los datos de cada producto
            for ingrediente in producto['ingredients']:
                ws.append([
                    producto['name'],
                    producto['description'],
                    producto['date'],
                    producto['yield_quantity'],
                    producto['total_cost'],
                    producto['cost_per_unit'],
                    ingrediente['name'],
                    # Puedes continuar añadiendo los detalles de cada ingrediente aquí
                ])
                
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        # Añade aquí la lógica para añadir los datos de `datos` al libro de Excel
        response = HttpResponse(virtual_workbook.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="recetario-Bakery-Merry.xlsx"'
        return response
    elif format == 'json':
        return JsonResponse(datos, safe=False)
    
    else:
         return JsonResponse({'error': 'Formato no soportado'}, status=400)